from configparser import ConfigParser
from datetime import datetime
from os import makedirs
from os.path import basename, isdir, isfile, join
from sys import getsizeof
from urllib.parse import quote
from uuid import uuid4

from openpyxl import Workbook, load_workbook


def configParserToDict(path):
    config = ConfigParser()
    config.read(path, encoding="utf-8")
    d = dict(config._sections)
    for k in d:
        d[k] = dict(d[k])
    return d




def flowIsLowLevel(threshold, curFlow, name):
    try:
        level = threshold[name.split("-")[0]][datetime.now().hour]
    except KeyError:
        return False, 0
    else:
        if curFlow < level:
            return True, level
        else:
            return False, level


def getLogTime():
    return datetime.now().strftime("[%m-%d %H:%M:%S]")


def saveCurrentFlow(r, path, name):
    if not isdir(path):
        makedirs(path)
    now = datetime.now()
    t = name.split('-')
    if len(t) > 1:
        name = "%s-%s.%s-%s.xlsx" % ("".join(t[:-1]), str(now.month), str(now.day), t[-1])
    else:
        name = "%s-%s.%s.xlsx" % (name, str(now.month), str(now.day))
    file = join(path, name)
    r.insert(0, "%s:%s" % (now.hour, now.minute))

    if isfile(file):
        book = load_workbook(file)
        sheet = book.get_sheet_by_name(book.get_sheet_names()[0])
    else:
        book = Workbook()
        sheet = book.get_sheet_by_name(book.get_sheet_names()[0])
        sheet.append(["时间", "用户带宽", "源站带宽", "用户连接数", "源站连接数"])

    sheet.append(r)
    try:
        book.save(file)
    except PermissionError:
        pass
    book.close()


def genNameFromPath(f_path, app):
    file = basename(f_path)
    name = "".join(file.split(".")[:-1])
    return name + '-' + str(uuid4()).split('-')[0]


def getFloat(v=None):
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0


def calStatistics(data, isRound=False):
    a = b = c = d = 0
    for i in data["tableData"]:
        a += getFloat(i[2])
        b += getFloat(i[3])
        c += getFloat(i[4])
        d += getFloat(i[5])
    if isRound:
        return [round(a), round(b), round(c), round(d)]
    return [a, b, c, d]


def parseLoginFile(info):
    ft = info["path"].split(".")[-1].lower()
    info["fileType"] = ft
    info["node"] = []
    info["tableData"] = []

    if ft == "txt":
        try:
            f = open(info["path"], encoding="utf-8")
            data = f.readlines()
        except UnicodeDecodeError:
            f = open(info["path"], encoding="gbk")
            data = f.readlines()

        for line in data:
            line = line.strip().split("----")
            if len(line) < 4:
                continue
            node = {
                "SessionID": None,
                "slogin": None,
                "flogin": (quote(line[1]), quote(line[2]), quote(line[3])),
                "loginCount": 0,
                "gotStatus": False,
                "needRefresh": False,
                "getStatusErrorNum": 0
            }
            info["node"].append(node)
            info["tableData"].append([line[0], line[1]] + [None for i in range(9)])
    else:
        f = load_workbook(info["path"], )
        sheet = f.worksheets[0]
        for row in sheet.rows:
            line = [str(each.value) for each in row]
            if len(line) < 4:
                continue
            node = {
                "SessionID": None,
                "slogin": None,
                "flogin": (quote(line[1]), quote(line[2]), quote(line[3])),
                "loginCount": 0,
                "gotStatus": False,
                "needRefresh": False,
                "getStatusErrorNum": 0
            }
            info["node"].append(node)
            info["tableData"].append([line[0], line[1]] + [None for i in range(9)])
    return info


def get_size(obj, seen=None):
    size = getsizeof(obj)
    if seen is None:
        seen = set()
    obj_id = id(obj)
    if obj_id in seen:
        return 0

    seen.add(obj_id)
    if isinstance(obj, dict):
        size += sum([get_size(v, seen) for v in obj.values()])
        size += sum([get_size(k, seen) for k in obj.keys()])
    elif hasattr(obj, '__dict__'):
        size += get_size(obj.__dict__, seen)
    elif hasattr(obj, '__iter__') and not isinstance(obj, (str, bytes, bytearray)):
        size += sum([get_size(i, seen) for i in obj])
    return size


if __name__ == '__main__':
    pass
